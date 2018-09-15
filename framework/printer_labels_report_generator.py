from openpyxl import load_workbook
from openpyxl import Workbook

class PrinterLabelsReportGenerator:
    def __init__(self, name):
        self._name = name

    def addRow(self, deviceId):
        try:
            wb = load_workbook(self._name)
        except FileNotFoundError:
            wb = Workbook()
            wb.active.append(['nr', 'DID', 'DID'])

        page = wb.active
        row = [str(page.max_row).zfill(4), deviceId, deviceId]
        page.append(row)
        wb.save(filename=self._name)